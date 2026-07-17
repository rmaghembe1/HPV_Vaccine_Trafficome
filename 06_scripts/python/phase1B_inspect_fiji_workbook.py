#!/usr/bin/env python3

from __future__ import annotations

import csv
import hashlib
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "ERROR: openpyxl is not installed.\n"
        "Install it using:\n"
        "  python -m pip install --user openpyxl"
    )


ROOT = Path("/mnt/d/HPV_Vaccine_Trafficome_Project")

WORKBOOK = (
    ROOT
    / "03_data_raw"
    / "hpv_specific"
    / "fiji_nct02276521"
    / "NCOMMS-24-64334A_HPV_collated_antibody_feature_data.xlsx"
)

OUTDIR = (
    ROOT
    / "02_dataset_audit"
    / "hpv_specific"
    / "fiji_nct02276521"
)

TABLEDIR = ROOT / "08_results" / "tables"
PREVIEW_DIR = OUTDIR / "sheet_previews"

EXPECTED_MD5 = "e42173e1d8297cd64420fd9682c42674"

SEARCH_TERMS = {
    "participant_identifier": [
        r"\bparticipant\b",
        r"\bsubject\b",
        r"\bsample[ _-]?id\b",
        r"\bparticipant[ _-]?id\b",
        r"\bsubject[ _-]?id\b",
        r"\bstudy[ _-]?id\b",
    ],
    "age": [
        r"\bage\b",
        r"\bage[ _-]?years?\b",
        r"\byears?[ _-]?old\b",
    ],
    "weight": [
        r"\bweight\b",
        r"\bbody[ _-]?weight\b",
        r"\bkilograms?\b",
        r"\bkg\b",
    ],
    "bmi_adiposity": [
        r"\bbmi\b",
        r"\bbody[ _-]?mass[ _-]?index\b",
        r"\badiposity\b",
        r"\bobes",
        r"\boverweight\b",
    ],
    "sex_gender": [
        r"\bsex\b",
        r"\bgender\b",
        r"\bfemale\b",
        r"\bmale\b",
    ],
    "ethnicity": [
        r"\bethnicity\b",
        r"\bethnic\b",
        r"\bi[- ]?taukei\b",
        r"\bindian\b",
        r"\bfijian\b",
    ],
    "dose_regimen": [
        r"\bdose\b",
        r"\bdosage\b",
        r"\bvaccination[ _-]?group\b",
        r"\bvaccine[ _-]?group\b",
        r"\bprevious[ _-]?dose",
    ],
    "time_visit": [
        r"\btime\b",
        r"\btimepoint\b",
        r"\bvisit\b",
        r"\bday\b",
        r"\bmonth\b",
        r"\byear\b",
        r"\bpre\b",
        r"\bpost\b",
        r"\bbaseline\b",
    ],
    "vaccine_type": [
        r"\bvaccine\b",
        r"\bquadrivalent\b",
        r"\bbivalent\b",
        r"\b4vhpv\b",
        r"\b2vhpv\b",
        r"\bgardasil\b",
        r"\bcervarix\b",
    ],
    "hpv_genotype": [
        r"\bhpv[ _-]?\d+",
        r"\bgenotype\b",
        r"\btype[ _-]?\d+",
    ],
    "antibody_isotype_subclass": [
        r"\bigg\b",
        r"\bigg1\b",
        r"\bigg2\b",
        r"\bigg3\b",
        r"\bigg4\b",
        r"\biga1\b",
        r"\biga2\b",
        r"\bigm\b",
    ],
    "fc_receptor": [
        r"\bfc.?r",
        r"\bfcgamma",
        r"\bfcγr",
    ],
    "adcp": [
        r"\badcp\b",
        r"\bphagocyt",
    ],
}

COMPILED_TERMS = {
    category: [re.compile(pattern, re.IGNORECASE) for pattern in patterns]
    for category, patterns in SEARCH_TERMS.items()
}


def compute_md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return cleaned or "sheet"


def find_categories(text: str) -> list[str]:
    categories: list[str] = []
    for category, patterns in COMPILED_TERMS.items():
        if any(pattern.search(text) for pattern in patterns):
            categories.append(category)
    return categories


def likely_header_row(preview_rows: list[list[Any]]) -> int | None:
    best_row = None
    best_score = -1

    for row_number, row in enumerate(preview_rows, start=1):
        text_values = [
            clean_text(value)
            for value in row
            if clean_text(value)
        ]

        if not text_values:
            continue

        string_count = sum(
            isinstance(value, str)
            for value in row
            if value is not None
        )

        matched_categories = set()
        for value in text_values:
            matched_categories.update(find_categories(value))

        score = (
            len(text_values)
            + 2 * string_count
            + 4 * len(matched_categories)
        )

        if score > best_score:
            best_score = score
            best_row = row_number

    return best_row


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fieldnames,
            delimiter="\t",
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    if not WORKBOOK.exists():
        sys.exit(f"ERROR: Workbook not found: {WORKBOOK}")

    OUTDIR.mkdir(parents=True, exist_ok=True)
    TABLEDIR.mkdir(parents=True, exist_ok=True)
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)

    observed_md5 = compute_md5(WORKBOOK)

    workbook = load_workbook(
        filename=WORKBOOK,
        read_only=True,
        data_only=True,
    )

    sheet_inventory: list[dict[str, Any]] = []
    metadata_hits: list[dict[str, Any]] = []
    candidate_headers: list[dict[str, Any]] = []

    category_counts: Counter[str] = Counter()
    category_sheets: defaultdict[str, set[str]] = defaultdict(set)

    for worksheet in workbook.worksheets:
        preview_rows: list[list[Any]] = []

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(20, worksheet.max_row),
            values_only=True,
        ):
            preview_rows.append(list(row))

        header_row = likely_header_row(preview_rows)

        nonempty_preview_cells = sum(
            1
            for row in preview_rows
            for value in row
            if clean_text(value)
        )

        sheet_inventory.append(
            {
                "sheet_name": worksheet.title,
                "max_rows": worksheet.max_row,
                "max_columns": worksheet.max_column,
                "likely_header_row": header_row or "",
                "nonempty_cells_first_20_rows": nonempty_preview_cells,
            }
        )

        preview_path = (
            PREVIEW_DIR
            / f"{safe_filename(worksheet.title)}_first20rows.tsv"
        )

        with preview_path.open(
            "w",
            encoding="utf-8",
            newline="",
        ) as handle:
            writer = csv.writer(handle, delimiter="\t")
            for row in preview_rows:
                writer.writerow(
                    [clean_text(value) for value in row]
                )

        if header_row is not None and header_row <= len(preview_rows):
            header_values = preview_rows[header_row - 1]

            for column_number, value in enumerate(
                header_values,
                start=1,
            ):
                text = clean_text(value)
                if not text:
                    continue

                categories = find_categories(text)

                candidate_headers.append(
                    {
                        "sheet_name": worksheet.title,
                        "header_row": header_row,
                        "column_number": column_number,
                        "header_value": text,
                        "matched_categories": ";".join(categories),
                    }
                )

        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            for column_number, value in enumerate(row, start=1):
                text = clean_text(value)

                if not text:
                    continue

                categories = find_categories(text)

                for category in categories:
                    category_counts[category] += 1
                    category_sheets[category].add(worksheet.title)

                    metadata_hits.append(
                        {
                            "category": category,
                            "sheet_name": worksheet.title,
                            "row_number": row_number,
                            "column_number": column_number,
                            "cell_value": text[:500],
                            "within_first_20_rows": (
                                "yes" if row_number <= 20 else "no"
                            ),
                        }
                    )

    workbook.close()

    inventory_path = (
        TABLEDIR
        / "phase1B_fiji_workbook_sheet_inventory.tsv"
    )

    hits_path = (
        TABLEDIR
        / "phase1B_fiji_candidate_metadata_hits.tsv"
    )

    headers_path = (
        TABLEDIR
        / "phase1B_fiji_candidate_headers.tsv"
    )

    write_tsv(
        inventory_path,
        [
            "sheet_name",
            "max_rows",
            "max_columns",
            "likely_header_row",
            "nonempty_cells_first_20_rows",
        ],
        sheet_inventory,
    )

    write_tsv(
        hits_path,
        [
            "category",
            "sheet_name",
            "row_number",
            "column_number",
            "cell_value",
            "within_first_20_rows",
        ],
        metadata_hits,
    )

    write_tsv(
        headers_path,
        [
            "sheet_name",
            "header_row",
            "column_number",
            "header_value",
            "matched_categories",
        ],
        candidate_headers,
    )

    report_path = (
        OUTDIR
        / "phase1B_fiji_workbook_inspection_report.md"
    )

    with report_path.open("w", encoding="utf-8") as report:
        report.write("# Phase 1B Fiji workbook inspection\n\n")

        report.write("## File integrity\n\n")
        report.write(f"- Workbook: `{WORKBOOK}`\n")
        report.write(f"- Expected MD5: `{EXPECTED_MD5}`\n")
        report.write(f"- Observed MD5: `{observed_md5}`\n")
        report.write(
            "- Integrity decision: "
            + (
                "**PASS**\n\n"
                if observed_md5 == EXPECTED_MD5
                else "**FAIL**\n\n"
            )
        )

        report.write("## Workbook structure\n\n")
        report.write(
            f"- Number of worksheets: {len(sheet_inventory)}\n"
        )

        report.write(
            f"- Total worksheet rows: "
            f"{sum(int(row['max_rows']) for row in sheet_inventory)}\n"
        )

        report.write(
            f"- Total detected keyword hits: "
            f"{len(metadata_hits)}\n\n"
        )

        report.write("## Worksheet inventory\n\n")
        report.write(
            "| Worksheet | Rows | Columns | Candidate header row |\n"
        )
        report.write(
            "|---|---:|---:|---:|\n"
        )

        for row in sheet_inventory:
            report.write(
                f"| {row['sheet_name']} "
                f"| {row['max_rows']} "
                f"| {row['max_columns']} "
                f"| {row['likely_header_row'] or 'not resolved'} |\n"
            )

        report.write("\n## Candidate variable categories\n\n")
        report.write(
            "| Category | Cell hits | Worksheets | Preliminary status |\n"
        )
        report.write(
            "|---|---:|---|---|\n"
        )

        for category in SEARCH_TERMS:
            count = category_counts.get(category, 0)
            sheets = sorted(category_sheets.get(category, set()))

            if count == 0:
                status = "Not detected"
            elif category == "participant_identifier":
                status = "Candidate identifier field detected"
            else:
                status = (
                    "Detected; participant-level availability "
                    "requires structural confirmation"
                )

            report.write(
                f"| {category} "
                f"| {count} "
                f"| {', '.join(sheets) if sheets else '—'} "
                f"| {status} |\n"
            )

        report.write("\n## Critical feasibility decisions\n\n")

        for category, label in [
            ("age", "Age"),
            ("weight", "Weight"),
            ("bmi_adiposity", "BMI/adiposity"),
            ("sex_gender", "Sex/gender"),
            ("ethnicity", "Ethnicity"),
            ("dose_regimen", "Dose/regimen"),
            ("time_visit", "Time/visit"),
        ]:
            count = category_counts.get(category, 0)

            if count:
                report.write(
                    f"- **{label}:** candidate workbook content detected; "
                    "inspect the candidate header and hit tables before "
                    "classifying it as participant-level metadata.\n"
                )
            else:
                report.write(
                    f"- **{label}:** no matching workbook field detected "
                    "by the first-pass lexical audit.\n"
                )

        report.write("\n## Generated outputs\n\n")
        report.write(f"- `{inventory_path}`\n")
        report.write(f"- `{headers_path}`\n")
        report.write(f"- `{hits_path}`\n")
        report.write(f"- `{PREVIEW_DIR}`\n")

        report.write("\n## Interpretation rule\n\n")
        report.write(
            "A keyword hit is not automatically a usable covariate. "
            "A variable will be classified as participant-level only after "
            "confirming that it is linked to a stable participant identifier, "
            "has sufficient nonmissing values, and is not merely part of a "
            "worksheet title, legend, assay label or explanatory note.\n"
        )

    print("===== PHASE 1B WORKBOOK INSPECTION COMPLETE =====")
    print(f"Workbook MD5: {observed_md5}")
    print(f"Integrity: {'PASS' if observed_md5 == EXPECTED_MD5 else 'FAIL'}")
    print(f"Worksheets: {len(sheet_inventory)}")
    print(f"Candidate metadata hits: {len(metadata_hits)}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
