#!/usr/bin/env python3

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/mnt/d/HPV_Vaccine_Trafficome_Project")

WORKBOOK = (
    ROOT
    / "03_data_raw"
    / "hpv_specific"
    / "fiji_nct02276521"
    / "NCOMMS-24-64334A_HPV_collated_antibody_feature_data.xlsx"
)

OUTDIR = ROOT / "08_results" / "tables"

VARIABLE_PATTERNS = {
    "participant_id": [
        r"participant",
        r"subject",
        r"sample.?id",
        r"study.?id",
        r"patient.?id",
        r"volunteer",
    ],
    "age": [
        r"^age$",
        r"age.?year",
        r"years?.?old",
    ],
    "weight": [
        r"^weight$",
        r"body.?weight",
        r"weight.?kg",
    ],
    "bmi": [
        r"^bmi$",
        r"body.?mass.?index",
        r"adiposity",
        r"obesity",
        r"overweight",
    ],
    "sex": [
        r"^sex$",
        r"^gender$",
        r"female",
        r"male",
    ],
    "ethnicity": [
        r"ethnicity",
        r"ethnic",
        r"i.?taukei",
        r"indo.?fijian",
    ],
    "dose": [
        r"dose",
        r"vaccination.?group",
        r"previous.?vaccin",
        r"number.?of.?dose",
    ],
    "visit_time": [
        r"visit",
        r"timepoint",
        r"baseline",
        r"pre.?boost",
        r"post.?boost",
        r"month",
        r"year",
    ],
    "vaccine": [
        r"vaccine",
        r"gardasil",
        r"cervarix",
        r"quadrivalent",
        r"bivalent",
    ],
}

COMPILED = {
    variable: [re.compile(pattern, re.IGNORECASE) for pattern in patterns]
    for variable, patterns in VARIABLE_PATTERNS.items()
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def variable_matches(value: str) -> list[str]:
    matched = []

    for variable, patterns in COMPILED.items():
        if any(pattern.search(value) for pattern in patterns):
            matched.append(variable)

    return matched


def is_numeric_or_category(value: Any) -> bool:
    if value is None or text(value) == "":
        return False

    if isinstance(value, (int, float)):
        return True

    normalized = text(value).lower()

    common_categories = {
        "male",
        "female",
        "m",
        "f",
        "yes",
        "no",
        "pre",
        "post",
        "baseline",
        "visit 1",
        "visit 2",
        "v1",
        "v2",
    }

    if normalized in common_categories:
        return True

    try:
        float(normalized)
        return True
    except ValueError:
        return False


def detect_header_rows(ws, max_scan_rows: int = 20) -> list[int]:
    candidate_rows = []

    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [
            text(ws.cell(row=row_number, column=col).value)
            for col in range(1, ws.max_column + 1)
        ]

        nonempty = [value for value in values if value]
        matches = sum(
            1
            for value in nonempty
            if variable_matches(value)
        )

        if len(nonempty) >= 2 and matches >= 1:
            candidate_rows.append(row_number)

    return candidate_rows


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    OUTDIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(
        WORKBOOK,
        read_only=True,
        data_only=True,
    )

    evidence_rows = []
    variable_summary = defaultdict(
        lambda: {
            "header_hits": 0,
            "participant_like_columns": 0,
            "glossary_hits": 0,
            "other_text_hits": 0,
            "sheets": set(),
            "examples": [],
        }
    )

    for ws in wb.worksheets:
        header_rows = detect_header_rows(ws)
        is_glossary = "glossary" in ws.title.lower()

        for row_number in range(1, min(ws.max_row, 30) + 1):
            for column_number in range(1, ws.max_column + 1):
                value = text(
                    ws.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                )

                if not value:
                    continue

                matched_variables = variable_matches(value)

                for variable in matched_variables:
                    summary = variable_summary[variable]
                    summary["sheets"].add(ws.title)

                    if len(summary["examples"]) < 8:
                        summary["examples"].append(
                            f"{ws.title}!R{row_number}C{column_number}={value}"
                        )

                    location_type = "other_text"

                    if is_glossary:
                        summary["glossary_hits"] += 1
                        location_type = "glossary"

                    elif row_number in header_rows:
                        summary["header_hits"] += 1
                        location_type = "candidate_header"

                        values_below = []

                        for downstream_row in range(
                            row_number + 1,
                            min(ws.max_row, row_number + 25) + 1,
                        ):
                            downstream_value = ws.cell(
                                row=downstream_row,
                                column=column_number,
                            ).value

                            if text(downstream_value):
                                values_below.append(downstream_value)

                        structured_values = sum(
                            is_numeric_or_category(item)
                            for item in values_below
                        )

                        if len(values_below) >= 3 and structured_values >= 3:
                            summary["participant_like_columns"] += 1
                            location_type = "structured_column"

                    else:
                        summary["other_text_hits"] += 1

                    evidence_rows.append(
                        {
                            "variable": variable,
                            "sheet": ws.title,
                            "row": row_number,
                            "column": column_number,
                            "matched_value": value,
                            "location_type": location_type,
                        }
                    )

    wb.close()

    summary_rows = []

    for variable in VARIABLE_PATTERNS:
        evidence = variable_summary[variable]

        if evidence["participant_like_columns"] > 0:
            decision = "PARTICIPANT_LEVEL_CANDIDATE"

        elif evidence["header_hits"] > 0:
            decision = "GROUP_LEVEL_OR_UNRESOLVED_HEADER"

        elif evidence["glossary_hits"] > 0:
            decision = "TEXT_OR_GLOSSARY_ONLY"

        elif evidence["other_text_hits"] > 0:
            decision = "TEXT_REFERENCE_ONLY"

        else:
            decision = "NOT_DETECTED"

        summary_rows.append(
            {
                "variable": variable,
                "decision": decision,
                "participant_like_columns": evidence[
                    "participant_like_columns"
                ],
                "header_hits": evidence["header_hits"],
                "glossary_hits": evidence["glossary_hits"],
                "other_text_hits": evidence["other_text_hits"],
                "worksheets": ";".join(sorted(evidence["sheets"])),
                "examples": " | ".join(evidence["examples"]),
            }
        )

    evidence_path = (
        OUTDIR
        / "phase1B2_fiji_metadata_structural_evidence.tsv"
    )

    summary_path = (
        OUTDIR
        / "phase1B2_fiji_metadata_decision.tsv"
    )

    with evidence_path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "variable",
                "sheet",
                "row",
                "column",
                "matched_value",
                "location_type",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(evidence_rows)

    with summary_path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "variable",
                "decision",
                "participant_like_columns",
                "header_hits",
                "glossary_hits",
                "other_text_hits",
                "worksheets",
                "examples",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    decision_counts = Counter(
        row["decision"]
        for row in summary_rows
    )

    report_path = (
        ROOT
        / "02_dataset_audit"
        / "hpv_specific"
        / "fiji_nct02276521"
        / "phase1B2_metadata_structure_decision.md"
    )

    with report_path.open("w", encoding="utf-8") as report:
        report.write("# Phase 1B2 metadata structure decision\n\n")

        report.write("## Decision summary\n\n")
        report.write("| Variable | Decision | Structured columns | Worksheets |\n")
        report.write("|---|---|---:|---|\n")

        for row in summary_rows:
            report.write(
                f"| {row['variable']} "
                f"| {row['decision']} "
                f"| {row['participant_like_columns']} "
                f"| {row['worksheets'] or '—'} |\n"
            )

        report.write("\n## Decision counts\n\n")

        for decision, count in sorted(decision_counts.items()):
            report.write(f"- {decision}: {count}\n")

        report.write("\n## Interpretation\n\n")
        report.write(
            "`PARTICIPANT_LEVEL_CANDIDATE` means that the workbook appears "
            "to contain a structured column with repeated values below a "
            "candidate header. It still requires confirmation that the field "
            "is linked to a stable participant identifier and is not an assay "
            "or group label.\n"
        )

        report.write(
            "\nVariables reported in the publication but absent from the "
            "workbook should be classified as requiring external linkage or "
            "author-provided participant metadata.\n"
        )

    print("===== PHASE 1B2 COMPLETE =====")
    print(f"Decision table: {summary_path}")
    print(f"Evidence table: {evidence_path}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
