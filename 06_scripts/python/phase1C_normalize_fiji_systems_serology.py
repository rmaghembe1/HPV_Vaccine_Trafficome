#!/usr/bin/env python3

from __future__ import annotations

import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/mnt/d/HPV_Vaccine_Trafficome_Project")

WORKBOOK = (
    ROOT
    / "03_data_raw"
    / "hpv_specific"
    / "fiji_nct02276521"
    / "NCOMMS-24-64334A_HPV_collated_antibody_feature_data.xlsx"
)

PROCESSED_DIR = (
    ROOT
    / "07_data_processed"
    / "fiji_nct02276521"
)

TABLE_DIR = ROOT / "08_results" / "tables"

AUDIT_DIR = (
    ROOT
    / "02_dataset_audit"
    / "hpv_specific"
    / "fiji_nct02276521"
)

SHEET_PATTERN = re.compile(
    r"^(HPV\d+|BPV)_(v1|v2)$",
    flags=re.IGNORECASE,
)

METADATA_COLUMNS = [
    "participant_id",
    "previous_4vHPV_doses",
    "visit",
    "visit_label",
    "study_time_description",
    "antigen_target",
    "antigen_class",
    "prior_vaccine_context",
    "booster_context",
    "source_sheet",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_identifier(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    value_text = clean_text(value)

    if re.fullmatch(r"\d+\.0", value_text):
        return value_text[:-2]

    return value_text


def normalize_dosage(value: Any) -> str:
    value_text = normalize_identifier(value)

    if not value_text:
        return ""

    try:
        numeric = float(value_text)

        if numeric.is_integer():
            return str(int(numeric))

        return str(numeric)

    except ValueError:
        return value_text


def write_tsv(
    path: Path,
    fieldnames: list[str],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fieldnames,
            delimiter="\t",
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)


def antigen_class(target: str) -> str:
    if target in {"HPV16", "HPV18"}:
        return "vaccine_target_2vHPV_and_4vHPV"

    if target in {"HPV31", "HPV33", "HPV45", "HPV52", "HPV58"}:
        return "cross_reactive_non_booster_target"

    if target == "BPV":
        return "heterologous_control_antigen"

    return "unclassified"


def assay_family(feature: str) -> str:
    normalized = feature.lower()

    if normalized in {
        "igg",
        "igm",
        "iga1",
        "iga2",
        "igg1",
        "igg2",
        "igg3",
        "igg4",
    }:
        return "antibody_isotype_or_subclass"

    if normalized.startswith("fcgr"):
        return "fc_gamma_receptor_binding"

    if normalized == "adcp":
        return "antibody_dependent_cellular_phagocytosis"

    if normalized == "nab":
        return "neutralizing_antibody"

    return "other_assay"


def main() -> None:
    if not WORKBOOK.exists():
        sys.exit(f"ERROR: Workbook not found: {WORKBOOK}")

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(
        WORKBOOK,
        read_only=True,
        data_only=True,
    )

    wide_rows: list[dict[str, Any]] = []
    long_rows: list[dict[str, Any]] = []

    all_features: list[str] = []
    feature_seen: set[str] = set()

    participant_dosages: defaultdict[str, set[str]] = defaultdict(set)
    participant_sheets: defaultdict[str, set[str]] = defaultdict(set)

    sheet_participants: defaultdict[
        tuple[str, str],
        set[str],
    ] = defaultdict(set)

    feature_availability: defaultdict[
        tuple[str, str, str],
        dict[str, int],
    ] = defaultdict(
        lambda: {
            "rows_total": 0,
            "nonmissing": 0,
        }
    )

    key_counts: Counter[tuple[str, str, str]] = Counter()
    sheet_summary_rows: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        match = SHEET_PATTERN.match(worksheet.title)

        if not match:
            continue

        target = match.group(1).upper()
        visit = match.group(2).lower()

        headers = [
            clean_text(cell.value)
            for cell in worksheet[1]
        ]

        if "ID" not in headers:
            raise ValueError(
                f"Required ID column absent from {worksheet.title}: "
                f"{headers}"
            )

        if "Dosage" not in headers:
            raise ValueError(
                f"Required Dosage column absent from {worksheet.title}: "
                f"{headers}"
            )

        if len(headers) != len(set(headers)):
            raise ValueError(
                f"Duplicate headers detected in {worksheet.title}: "
                f"{headers}"
            )

        id_index = headers.index("ID")
        dosage_index = headers.index("Dosage")

        features = [
            header
            for header in headers
            if header not in {"ID", "Dosage"} and header
        ]

        for feature in features:
            if feature not in feature_seen:
                all_features.append(feature)
                feature_seen.add(feature)

        sheet_row_count = 0
        sheet_ids: set[str] = set()

        for values in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            participant_id = normalize_identifier(
                values[id_index]
                if id_index < len(values)
                else None
            )

            if not participant_id:
                continue

            dosage = normalize_dosage(
                values[dosage_index]
                if dosage_index < len(values)
                else None
            )

            if visit == "v1":
                visit_label = "persistence_pre_booster"
                study_time_description = (
                    "Six years after the participant's last 4vHPV dose; "
                    "before the study 2vHPV booster"
                )
                booster_context = "pre_2vHPV_booster"
            else:
                visit_label = "recall_post_booster_day28"
                study_time_description = (
                    "Twenty-eight days after the study 2vHPV booster"
                )
                booster_context = "post_2vHPV_booster_day28"

            record: dict[str, Any] = {
                "participant_id": participant_id,
                "previous_4vHPV_doses": dosage,
                "visit": visit,
                "visit_label": visit_label,
                "study_time_description": study_time_description,
                "antigen_target": target,
                "antigen_class": antigen_class(target),
                "prior_vaccine_context": (
                    f"{dosage}_previous_4vHPV_doses"
                    if dosage
                    else "previous_4vHPV_doses_unknown"
                ),
                "booster_context": booster_context,
                "source_sheet": worksheet.title,
            }

            for feature in all_features:
                record[feature] = ""

            for feature in features:
                feature_index = headers.index(feature)

                value = (
                    values[feature_index]
                    if feature_index < len(values)
                    else None
                )

                record[feature] = value if value is not None else ""

                availability_key = (
                    target,
                    visit,
                    feature,
                )

                feature_availability[availability_key][
                    "rows_total"
                ] += 1

                if value is not None and clean_text(value):
                    feature_availability[availability_key][
                        "nonmissing"
                    ] += 1

                    long_rows.append(
                        {
                            **{
                                key: record[key]
                                for key in METADATA_COLUMNS
                            },
                            "feature": feature,
                            "assay_family": assay_family(feature),
                            "value": value,
                        }
                    )

            wide_rows.append(record)

            participant_dosages[participant_id].add(dosage)
            participant_sheets[participant_id].add(
                worksheet.title
            )

            sheet_participants[(target, visit)].add(
                participant_id
            )

            key_counts[
                (
                    participant_id,
                    target,
                    visit,
                )
            ] += 1

            sheet_ids.add(participant_id)
            sheet_row_count += 1

        sheet_summary_rows.append(
            {
                "source_sheet": worksheet.title,
                "antigen_target": target,
                "antigen_class": antigen_class(target),
                "visit": visit,
                "participant_rows": sheet_row_count,
                "unique_participants": len(sheet_ids),
                "assay_features": ";".join(features),
                "number_of_assay_features": len(features),
            }
        )

    workbook.close()

    # Ensure every previously constructed row has the full union of features.
    for record in wide_rows:
        for feature in all_features:
            record.setdefault(feature, "")

    duplicate_rows = [
        {
            "participant_id": key[0],
            "antigen_target": key[1],
            "visit": key[2],
            "row_count": count,
        }
        for key, count in sorted(key_counts.items())
        if count > 1
    ]

    dosage_inconsistency_rows = []

    for participant_id, values in sorted(
        participant_dosages.items()
    ):
        nonempty_values = sorted(
            value
            for value in values
            if value
        )

        if len(set(nonempty_values)) > 1:
            dosage_inconsistency_rows.append(
                {
                    "participant_id": participant_id,
                    "observed_dosages": ";".join(
                        nonempty_values
                    ),
                    "number_of_dosage_values": len(
                        set(nonempty_values)
                    ),
                }
            )

    participant_metadata_rows = []

    for participant_id in sorted(
        participant_dosages,
        key=lambda value: (
            not value.isdigit(),
            int(value) if value.isdigit() else value,
        ),
    ):
        dosage_values = sorted(
            value
            for value in participant_dosages[
                participant_id
            ]
            if value
        )

        participant_metadata_rows.append(
            {
                "participant_id": participant_id,
                "previous_4vHPV_doses": (
                    dosage_values[0]
                    if len(set(dosage_values)) == 1
                    else ";".join(sorted(set(dosage_values)))
                ),
                "dosage_consistent_across_sheets": (
                    "yes"
                    if len(set(dosage_values)) <= 1
                    else "no"
                ),
                "number_of_source_sheets": len(
                    participant_sheets[participant_id]
                ),
                "age_available": "no",
                "weight_available": "no",
                "bmi_available": "no",
                "sex_available": "no",
                "ethnicity_available": "no",
            }
        )

    pairing_rows = []

    targets = sorted(
        {
            target
            for target, visit in sheet_participants
        }
    )

    for target in targets:
        v1_ids = sheet_participants.get(
            (target, "v1"),
            set(),
        )

        v2_ids = sheet_participants.get(
            (target, "v2"),
            set(),
        )

        paired_ids = v1_ids & v2_ids
        v1_only_ids = v1_ids - v2_ids
        v2_only_ids = v2_ids - v1_ids

        pairing_rows.append(
            {
                "antigen_target": target,
                "antigen_class": antigen_class(target),
                "visit1_participants": len(v1_ids),
                "visit2_participants": len(v2_ids),
                "paired_participants": len(paired_ids),
                "visit1_only_participants": len(v1_only_ids),
                "visit2_only_participants": len(v2_only_ids),
                "paired_fraction_of_union": (
                    round(
                        len(paired_ids)
                        / len(v1_ids | v2_ids),
                        6,
                    )
                    if (v1_ids | v2_ids)
                    else ""
                ),
                "visit1_only_ids": ";".join(
                    sorted(v1_only_ids)
                ),
                "visit2_only_ids": ";".join(
                    sorted(v2_only_ids)
                ),
            }
        )

    feature_rows = []

    for (
        target,
        visit,
        feature,
    ), counts in sorted(feature_availability.items()):
        rows_total = counts["rows_total"]
        nonmissing = counts["nonmissing"]

        feature_rows.append(
            {
                "antigen_target": target,
                "antigen_class": antigen_class(target),
                "visit": visit,
                "feature": feature,
                "assay_family": assay_family(feature),
                "rows_total": rows_total,
                "nonmissing_values": nonmissing,
                "missing_values": rows_total - nonmissing,
                "nonmissing_fraction": (
                    round(nonmissing / rows_total, 6)
                    if rows_total
                    else ""
                ),
            }
        )

    coverage_map: defaultdict[
        tuple[str, str],
        dict[str, Any],
    ] = defaultdict(
        lambda: {
            "targets": set(),
            "observations": 0,
        }
    )

    for row in wide_rows:
        key = (
            row["participant_id"],
            row["visit"],
        )

        coverage_map[key]["targets"].add(
            row["antigen_target"]
        )

        coverage_map[key]["observations"] += sum(
            1
            for feature in all_features
            if row.get(feature, "") not in {"", None}
        )

    participant_coverage_rows = []

    for (
        participant_id,
        visit,
    ), information in sorted(coverage_map.items()):
        targets_present = sorted(information["targets"])

        participant_coverage_rows.append(
            {
                "participant_id": participant_id,
                "visit": visit,
                "number_of_antigen_targets": len(
                    targets_present
                ),
                "antigen_targets": ";".join(
                    targets_present
                ),
                "number_of_nonmissing_assay_values": (
                    information["observations"]
                ),
                "complete_eight_target_coverage": (
                    "yes"
                    if len(targets_present) == 8
                    else "no"
                ),
            }
        )

    wide_path = (
        PROCESSED_DIR
        / "phase1C_fiji_participant_antigen_visit_wide.tsv"
    )

    long_path = (
        PROCESSED_DIR
        / "phase1C_fiji_participant_antigen_visit_feature_long.tsv"
    )

    participant_metadata_path = (
        PROCESSED_DIR
        / "phase1C_fiji_participant_metadata.tsv"
    )

    write_tsv(
        wide_path,
        METADATA_COLUMNS + all_features,
        wide_rows,
    )

    write_tsv(
        long_path,
        METADATA_COLUMNS
        + [
            "feature",
            "assay_family",
            "value",
        ],
        long_rows,
    )

    write_tsv(
        participant_metadata_path,
        [
            "participant_id",
            "previous_4vHPV_doses",
            "dosage_consistent_across_sheets",
            "number_of_source_sheets",
            "age_available",
            "weight_available",
            "bmi_available",
            "sex_available",
            "ethnicity_available",
        ],
        participant_metadata_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_sheet_summary.tsv",
        [
            "source_sheet",
            "antigen_target",
            "antigen_class",
            "visit",
            "participant_rows",
            "unique_participants",
            "assay_features",
            "number_of_assay_features",
        ],
        sheet_summary_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_pairing_audit.tsv",
        [
            "antigen_target",
            "antigen_class",
            "visit1_participants",
            "visit2_participants",
            "paired_participants",
            "visit1_only_participants",
            "visit2_only_participants",
            "paired_fraction_of_union",
            "visit1_only_ids",
            "visit2_only_ids",
        ],
        pairing_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_feature_availability.tsv",
        [
            "antigen_target",
            "antigen_class",
            "visit",
            "feature",
            "assay_family",
            "rows_total",
            "nonmissing_values",
            "missing_values",
            "nonmissing_fraction",
        ],
        feature_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_participant_coverage.tsv",
        [
            "participant_id",
            "visit",
            "number_of_antigen_targets",
            "antigen_targets",
            "number_of_nonmissing_assay_values",
            "complete_eight_target_coverage",
        ],
        participant_coverage_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_duplicate_key_audit.tsv",
        [
            "participant_id",
            "antigen_target",
            "visit",
            "row_count",
        ],
        duplicate_rows,
    )

    write_tsv(
        TABLE_DIR / "phase1C_fiji_dosage_consistency_audit.tsv",
        [
            "participant_id",
            "observed_dosages",
            "number_of_dosage_values",
        ],
        dosage_inconsistency_rows,
    )

    corrected_decisions = [
        {
            "variable": "participant_id",
            "decision": "PARTICIPANT_LEVEL_CONFIRMED",
            "source": "ID column in all assay worksheets",
            "analysis_use": "pairing and participant random effects",
        },
        {
            "variable": "previous_4vHPV_doses",
            "decision": "PARTICIPANT_LEVEL_CONFIRMED",
            "source": "Dosage column in all assay worksheets",
            "analysis_use": "primary regimen exposure",
        },
        {
            "variable": "visit",
            "decision": "DERIVED_FROM_WORKSHEET",
            "source": "_v1 and _v2 worksheet suffixes",
            "analysis_use": "persistence versus booster recall",
        },
        {
            "variable": "antigen_target",
            "decision": "DERIVED_FROM_WORKSHEET",
            "source": "HPV16, HPV18, HPV31, HPV33, HPV45, HPV52, HPV58 or BPV worksheet prefix",
            "analysis_use": "vaccine-target, cross-reactive and control contrasts",
        },
        {
            "variable": "prior_vaccine_type",
            "decision": "STUDY_LEVEL_CONFIRMED",
            "source": "Glossary identifies previous doses as 4vHPV",
            "analysis_use": "prior exposure context",
        },
        {
            "variable": "booster_vaccine_type",
            "decision": "STUDY_LEVEL_CONFIRMED",
            "source": "Glossary identifies Visit 2 as 28 days after 2vHPV booster",
            "analysis_use": "booster-recall context",
        },
        {
            "variable": "age",
            "decision": "NOT_AVAILABLE_IN_OPEN_WORKBOOK",
            "source": "No participant-level field detected",
            "analysis_use": "requires external linkage",
        },
        {
            "variable": "weight",
            "decision": "NOT_AVAILABLE_IN_OPEN_WORKBOOK",
            "source": "No participant-level field detected",
            "analysis_use": "requires external linkage",
        },
        {
            "variable": "bmi",
            "decision": "NOT_AVAILABLE_IN_OPEN_WORKBOOK",
            "source": "No participant-level field detected",
            "analysis_use": "requires external linkage",
        },
        {
            "variable": "sex",
            "decision": "NOT_AVAILABLE_IN_OPEN_WORKBOOK",
            "source": "No participant-level field detected",
            "analysis_use": "not estimable as a covariate",
        },
        {
            "variable": "ethnicity",
            "decision": "NOT_AVAILABLE_IN_OPEN_WORKBOOK",
            "source": "No participant-level field detected",
            "analysis_use": "requires external linkage",
        },
    ]

    write_tsv(
        TABLE_DIR
        / "phase1B2_fiji_metadata_decision_corrected.tsv",
        [
            "variable",
            "decision",
            "source",
            "analysis_use",
        ],
        corrected_decisions,
    )

    report_path = (
        AUDIT_DIR
        / "phase1C_fiji_normalization_and_pairing_report.md"
    )

    unique_participants = len(participant_metadata_rows)

    dose_counts = Counter(
        row["previous_4vHPV_doses"]
        for row in participant_metadata_rows
    )

    with report_path.open(
        "w",
        encoding="utf-8",
    ) as report:
        report.write(
            "# Phase 1C Fiji normalization and pairing report\n\n"
        )

        report.write("## Structural decision\n\n")
        report.write(
            "The previous Phase 1B2 lexical classifier produced a "
            "false-negative decision because it did not recognize the "
            "exact workbook headers `ID` and `Dosage`, and because visit "
            "and antigen target are encoded in worksheet names. The "
            "Phase 1B2 report is superseded by the corrected decision "
            "table generated in this phase.\n\n"
        )

        report.write("## Normalized dataset\n\n")
        report.write(
            f"- Unique participants: {unique_participants}\n"
        )
        report.write(
            f"- Participant–antigen–visit rows: {len(wide_rows)}\n"
        )
        report.write(
            f"- Nonmissing feature-level observations: "
            f"{len(long_rows)}\n"
        )
        report.write(
            f"- Antigen targets: "
            f"{', '.join(sorted(targets))}\n"
        )
        report.write(
            f"- Assay features: "
            f"{', '.join(all_features)}\n"
        )
        report.write(
            f"- Duplicate participant–antigen–visit keys: "
            f"{len(duplicate_rows)}\n"
        )
        report.write(
            f"- Participants with inconsistent dosage values: "
            f"{len(dosage_inconsistency_rows)}\n\n"
        )

        report.write("## Previous 4vHPV dose distribution\n\n")
        report.write("| Previous doses | Participants |\n")
        report.write("|---:|---:|\n")

        for dose, count in sorted(
            dose_counts.items(),
            key=lambda item: item[0],
        ):
            report.write(f"| {dose or 'missing'} | {count} |\n")

        report.write("\n## Pairing by antigen\n\n")
        report.write(
            "| Antigen | Visit 1 | Visit 2 | Paired | V1 only | V2 only |\n"
        )
        report.write("|---|---:|---:|---:|---:|---:|\n")

        for row in pairing_rows:
            report.write(
                f"| {row['antigen_target']} "
                f"| {row['visit1_participants']} "
                f"| {row['visit2_participants']} "
                f"| {row['paired_participants']} "
                f"| {row['visit1_only_participants']} "
                f"| {row['visit2_only_participants']} |\n"
            )

        report.write("\n## Metadata feasibility\n\n")
        report.write(
            "- Participant ID: directly available.\n"
            "- Previous 4vHPV dose number: directly available.\n"
            "- Visit and booster timing: derivable from worksheet "
            "suffix and glossary.\n"
            "- Antigen target: derivable from worksheet name.\n"
            "- Age, weight, BMI, sex and ethnicity: absent from the "
            "open workbook.\n"
        )

        report.write("\n## Analysis decision\n\n")
        report.write(
            "The Fiji cohort is suitable for participant-paired analysis "
            "of long-term antibody persistence and day-28 booster recall "
            "by previous 4vHPV dose number, antigen target and antibody "
            "functional layer. It is not independently suitable for age, "
            "BMI, sex or ethnicity modeling without external participant "
            "metadata linkage.\n"
        )

        report.write("\n## Primary analytical contrasts\n\n")
        report.write(
            "1. Visit 2 versus Visit 1 within participants.\n"
            "2. Interaction between visit and previous 4vHPV dose number.\n"
            "3. HPV16/18 vaccine-target responses versus HPV31/33/45/52/58 "
            "cross-reactive responses.\n"
            "4. HPV-specific responses versus BPV heterologous control.\n"
            "5. Antibody quantity versus Fc-receptor engagement, ADCP and "
            "neutralization.\n"
        )

    print("===== PHASE 1C COMPLETE =====")
    print(f"Unique participants: {unique_participants}")
    print(f"Wide rows: {len(wide_rows)}")
    print(f"Long feature rows: {len(long_rows)}")
    print(f"Duplicate keys: {len(duplicate_rows)}")
    print(
        "Dosage inconsistencies: "
        f"{len(dosage_inconsistency_rows)}"
    )
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
